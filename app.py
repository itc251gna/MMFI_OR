import hashlib
import ipaddress
import json
import os
import shutil
import subprocess
import tempfile
import time
import uuid
from datetime import date, datetime, timedelta

import pyzipper
from apscheduler.schedulers.background import BackgroundScheduler
from dotenv import load_dotenv
from flask import (
    Flask,
    Response,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask_login import (
    LoginManager,
    UserMixin,
    current_user,
    login_required,
    login_user,
    logout_user,
)
from flask_sqlalchemy import SQLAlchemy
from flask_wtf.csrf import CSRFError, CSRFProtect
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.exc import IntegrityError
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

load_dotenv()

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "static", "uploads")
BACKUP_FOLDER = os.getenv("BACKUP_FOLDER") or os.path.join(BASE_DIR, "runtime", "backups")


def bool_env(name, default="0"):
    return os.getenv(name, default).strip().lower() in {"1", "true", "yes", "on"}


def csv_env(name, default=""):
    return [item.strip() for item in os.getenv(name, default).split(",") if item.strip()]


def required_env(name, fallback=None):
    value = os.getenv(name)
    if value:
        return value
    if fallback is not None:
        return fallback
    raise RuntimeError(f"Required environment variable {name} is not set")


WORKSTATIONS = [
    {
        "code": "surgery_secretariat",
        "name": "Γραμματεία Χειρουργικού Τομέα",
        "short": "Γραμματεία Χειρ.",
    },
    {
        "code": "isupply_committee",
        "name": "Επιτροπή I-Supply",
        "short": "I-Supply",
    },
    {
        "code": "orders_office",
        "name": "ΔΟΥ/Τμ. Προμ./Γρ. Παραγγελιών",
        "short": "Γρ. Παραγγελιών",
    },
    {
        "code": "mef",
        "name": "ΜΕΦ",
        "short": "ΜΕΦ",
    },
    {
        "code": "budget_accounting",
        "name": "ΔΟΥ/Τμ. Λογ./Π/Υ",
        "short": "Λογ./Π/Υ",
    },
    {
        "code": "small_procurement",
        "name": "ΔΟΥ/Τμ. Προμ./Γρ. Μικροπρομηθειών",
        "short": "Μικροπρομήθειες",
    },
    {
        "code": "receiving_committee",
        "name": "Επιτροπή Παραλαβών",
        "short": "Παραλαβές",
    },
    {
        "code": "finance_office",
        "name": "ΔΟΥ/Τμ. Οικονομικό",
        "short": "Οικονομικό",
    },
]

WORKSTATION_BY_CODE = {station["code"]: station for station in WORKSTATIONS}
ALL_STATION_CODES = [station["code"] for station in WORKSTATIONS]

WORKFLOW_DEFINITION = [
    {
        "order": 1,
        "station_code": "surgery_secretariat",
        "title": "Εγγραφή χειρουργείου στην εφαρμογή",
        "description": "Ημερομηνία χειρουργείου, περιγραφή, υπεύθυνος ιατρός, έρευνα αγοράς ή/και έκθεση αναγκαιότητας σκοπιμότητας.",
    },
    {
        "order": 2,
        "station_code": "isupply_committee",
        "title": "Έρευνα αγοράς και πρακτικό",
        "description": "Ολοκλήρωση ενεργειών Επιτροπής I-Supply.",
    },
    {
        "order": 3,
        "station_code": "orders_office",
        "title": "Τοποθέτηση παραγγελίας εκτός εφαρμογής SAP",
        "description": "Διαδικασία εκτός εφαρμογής SAP από το Γραφείο Παραγγελιών.",
    },
    {
        "order": 4,
        "station_code": "surgery_secretariat",
        "title": "Έκδοση Υ.Σ. μέσω Medico προς ΜΕΦ",
        "description": "Έκδοση υπηρεσιακού σημειώματος για αίτημα προμήθειας υλικών μέσω εφαρμογής SAP.",
    },
    {
        "order": 5,
        "station_code": "mef",
        "title": "Έκδοση υπόψη αιτήματος",
        "description": "Ενέργειες ΜΕΦ για έκδοση υπόψη αιτήματος.",
    },
    {
        "order": 6,
        "station_code": "budget_accounting",
        "title": "Χρηματοδότηση αιτήματος",
        "description": "Χρηματοδότηση από ΔΟΥ/Τμ. Λογ./Π/Υ.",
    },
    {
        "order": 7,
        "station_code": "small_procurement",
        "title": "Έκδοση ΠΥΠ, απόφαση ανάθεσης και αναρτήσεις",
        "description": "Ενέργειες Γραφείου Μικροπρομηθειών.",
    },
    {
        "order": 8,
        "station_code": "orders_office",
        "title": "Ιδιωτικό συμφωνητικό και αποστολή στην εταιρεία",
        "description": "Αποστολή για υπογραφή και στη συνέχεια τιμολόγηση.",
    },
    {
        "order": 9,
        "station_code": "orders_office",
        "title": "Παραλαβή τιμολογίων και αποστολή στην Επιτροπή Παραλαβών",
        "description": "Διαβίβαση τιμολογίων από Γραφείο Παραγγελιών.",
    },
    {
        "order": 10,
        "station_code": "receiving_committee",
        "title": "Παραλαβή τιμολογίων και αποστολή στη ΜΕΦ",
        "description": "Αποστολή στη ΜΕΦ για λογιστική τακτοποίηση.",
    },
    {
        "order": 11,
        "station_code": "mef",
        "title": "MIGO-MIRO, γραμμάτια και διαβίβαση στη ΔΟΥ",
        "description": "Λογιστική τακτοποίηση, έκδοση γραμματίων και διαβίβαση τιμολογίων στη ΔΟΥ.",
    },
    {
        "order": 12,
        "station_code": "orders_office",
        "title": "Παραλαβή από ΜΕΦ και διαβίβαση σε ΔΟΥ/Τμ. Οικ.",
        "description": "Παραλαβή τιμολογίων από ΜΕΦ και διαβίβαση στο Οικονομικό.",
    },
    {
        "order": 13,
        "station_code": "finance_office",
        "title": "Προετοιμασία αποστολής λογαριασμού προς ΕΛΔΑ/Α",
        "description": "Τελική προετοιμασία λογαριασμού από ΔΟΥ/Τμ. Οικονομικό.",
    },
]

CASE_STATUS_LABELS = {
    "open": "Σε εξέλιξη",
    "completed": "Ολοκληρωμένο",
    "cancelled": "Ακυρωμένο",
}

STEP_STATUS_LABELS = {
    "pending": "Εκκρεμεί",
    "completed": "Ολοκληρώθηκε",
}


app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1, x_prefix=1)
app.secret_key = os.getenv("FLASK_SECRET_KEY") or os.getenv("SECRET_KEY") or "dev-only-change-me"
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=int(os.getenv("SESSION_HOURS", "8")))
app.config["SESSION_COOKIE_NAME"] = os.getenv("SESSION_COOKIE_NAME", "mmfi_session")
app.config["SESSION_COOKIE_SECURE"] = bool_env("SESSION_COOKIE_SECURE", "0")
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SQLALCHEMY_DATABASE_URI"] = os.getenv("DATABASE_URL") or "sqlite:///" + os.path.join(BASE_DIR, "runtime", "mmfi.sqlite3")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = int(os.getenv("MAX_UPLOAD_MB", "25")) * 1024 * 1024

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(BACKUP_FOLDER, exist_ok=True)
os.makedirs(os.path.join(BASE_DIR, "runtime"), exist_ok=True)

csrf = CSRFProtect(app)
db = SQLAlchemy(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"
login_manager.login_message = "Παρακαλώ συνδεθείτε για να συνεχίσετε."
login_manager.login_message_category = "warning"
login_manager.session_protection = "basic"

TRUST_SSO_HEADERS = bool_env("TRUST_SSO_HEADERS", "0")
SSO_TRUSTED_PROXY_CIDRS = csv_env(
    "SSO_TRUSTED_PROXY_CIDRS",
    "127.0.0.1/32,172.16.0.0/12,192.168.0.0/16",
)
SSO_APP_USER_GROUP = os.getenv("SSO_APP_USER_GROUP", "/apps/mmfi/users")
SSO_APP_ADMIN_GROUP = os.getenv("SSO_APP_ADMIN_GROUP", "/apps/mmfi/admins")
SSO_GLOBAL_ADMIN_GROUP = os.getenv("SSO_GLOBAL_ADMIN_GROUP", "/apps/global/admins")
SSO_STATION_GROUP_PREFIX = os.getenv("SSO_STATION_GROUP_PREFIX", "/apps/mmfi/stations").rstrip("/")
ALLOW_SSO_USER_GROUP_ALL_STATIONS = bool_env("ALLOW_SSO_USER_GROUP_ALL_STATIONS", "0")
CENTRAL_AUTH_REALM = os.getenv("CENTRAL_AUTH_REALM", "intranet")
CENTRAL_AUTH_ADMIN_URL = os.getenv("CENTRAL_AUTH_ADMIN_URL", "https://auth.251gh.local/admin/")
CENTRAL_AUTH_USERS_URL = os.getenv(
    "CENTRAL_AUTH_USERS_URL",
    f"https://auth.251gh.local/admin/master/console/#/{CENTRAL_AUTH_REALM}/users",
)
CENTRAL_AUTH_GROUPS_URL = os.getenv(
    "CENTRAL_AUTH_GROUPS_URL",
    f"https://auth.251gh.local/admin/master/console/#/{CENTRAL_AUTH_REALM}/groups",
)
ALLOW_LOCAL_USER_ADMIN_FROM_SSO = bool_env("ALLOW_LOCAL_USER_ADMIN_FROM_SSO", "0")
BACKUP_RETENTION_COUNT = int(os.getenv("BACKUP_RETENTION_COUNT", "30"))
BACKUP_INCLUDE_DATABASE = bool_env("BACKUP_INCLUDE_DATABASE", "1")
DELAY_WARNING_DAYS = int(os.getenv("DELAY_WARNING_DAYS", "10"))


class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(255), nullable=False)
    is_admin = db.Column(db.Boolean, default=False, nullable=False)
    is_active_local = db.Column(db.Boolean, default=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    @property
    def is_active(self):
        return self.is_active_local

    @property
    def auth_method(self):
        return "local"


class LocalUserStationPermission(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    station_code = db.Column(db.String(80), nullable=False, index=True)
    user = db.relationship("User", backref=db.backref("station_permissions", cascade="all, delete-orphan"))

    __table_args__ = (
        db.UniqueConstraint("user_id", "station_code", name="uq_local_user_station"),
    )


class SSOUser(UserMixin):
    def __init__(self, username, *, sso_admin=False, groups=None, email=None):
        self.id = f"sso:{username}"
        self.username = username
        self.email = email
        self.auth_method = "sso"
        self.sso_admin = bool(sso_admin)
        self.sso_groups = groups or []


class SurgeryCase(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    reference_code = db.Column(db.String(30), unique=True, nullable=False, index=True)
    surgery_date = db.Column(db.Date, nullable=False, index=True)
    surgery_description = db.Column(db.Text, nullable=False)
    responsible_doctor = db.Column(db.String(160), nullable=False)
    supplier = db.Column(db.String(180))
    internal_reference = db.Column(db.String(120))
    notes = db.Column(db.Text)
    priority = db.Column(db.String(20), default="normal", nullable=False)
    status = db.Column(db.String(20), default="open", nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)
    created_by = db.Column(db.String(120))
    completed_at = db.Column(db.DateTime)
    cancelled_at = db.Column(db.DateTime)
    cancelled_by = db.Column(db.String(120))
    cancellation_reason = db.Column(db.Text)

    steps = db.relationship(
        "WorkflowStep",
        back_populates="case",
        cascade="all, delete-orphan",
        order_by="WorkflowStep.step_order",
    )
    attachments = db.relationship(
        "CaseAttachment",
        back_populates="case",
        cascade="all, delete-orphan",
        order_by="CaseAttachment.uploaded_at.desc()",
    )


class WorkflowStep(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    case_id = db.Column(db.Integer, db.ForeignKey("surgery_case.id"), nullable=False, index=True)
    step_order = db.Column(db.Integer, nullable=False)
    station_code = db.Column(db.String(80), nullable=False, index=True)
    station_name = db.Column(db.String(180), nullable=False)
    action_title = db.Column(db.String(220), nullable=False)
    action_description = db.Column(db.Text, nullable=False)
    status = db.Column(db.String(20), default="pending", nullable=False, index=True)
    notes = db.Column(db.Text)
    completed_at = db.Column(db.DateTime)
    completed_by = db.Column(db.String(120))
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

    case = db.relationship("SurgeryCase", back_populates="steps")
    attachments = db.relationship("CaseAttachment", back_populates="step")
    history_entries = db.relationship(
        "StepHistoryEntry",
        back_populates="step",
        cascade="all, delete-orphan",
        order_by="StepHistoryEntry.version_number.desc()",
    )

    __table_args__ = (
        db.UniqueConstraint("case_id", "step_order", name="uq_case_step_order"),
    )


class CaseAttachment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    case_id = db.Column(db.Integer, db.ForeignKey("surgery_case.id"), nullable=False, index=True)
    step_id = db.Column(db.Integer, db.ForeignKey("workflow_step.id"), index=True)
    original_filename = db.Column(db.String(255), nullable=False)
    stored_filename = db.Column(db.String(255), nullable=False)
    relative_path = db.Column(db.String(500), nullable=False)
    description = db.Column(db.String(255))
    uploaded_by = db.Column(db.String(120))
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    case = db.relationship("SurgeryCase", back_populates="attachments")
    step = db.relationship("WorkflowStep", back_populates="attachments")


class StepHistoryEntry(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    case_id = db.Column(db.Integer, db.ForeignKey("surgery_case.id"), nullable=False, index=True)
    step_id = db.Column(db.Integer, db.ForeignKey("workflow_step.id"), nullable=False, index=True)
    version_number = db.Column(db.Integer, nullable=False)
    event_type = db.Column(db.String(30), default="update", nullable=False, index=True)
    comment = db.Column(db.Text)
    created_by = db.Column(db.String(120))
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)

    case = db.relationship("SurgeryCase", backref=db.backref("step_history_entries", cascade="all, delete-orphan"))
    step = db.relationship("WorkflowStep", back_populates="history_entries")
    files = db.relationship(
        "StepHistoryAttachment",
        back_populates="history_entry",
        cascade="all, delete-orphan",
        order_by="StepHistoryAttachment.uploaded_at.asc()",
    )

    __table_args__ = (
        db.UniqueConstraint("step_id", "version_number", name="uq_step_history_version"),
    )


class StepHistoryAttachment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    history_entry_id = db.Column(db.Integer, db.ForeignKey("step_history_entry.id"), nullable=False, index=True)
    case_id = db.Column(db.Integer, db.ForeignKey("surgery_case.id"), nullable=False, index=True)
    step_id = db.Column(db.Integer, db.ForeignKey("workflow_step.id"), nullable=False, index=True)
    original_filename = db.Column(db.String(255), nullable=False)
    stored_filename = db.Column(db.String(255), nullable=False)
    relative_path = db.Column(db.String(500), nullable=False)
    uploaded_by = db.Column(db.String(120))
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    history_entry = db.relationship("StepHistoryEntry", back_populates="files")
    case = db.relationship("SurgeryCase")
    step = db.relationship("WorkflowStep")


class AuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    username = db.Column(db.String(120))
    action = db.Column(db.String(80), nullable=False, index=True)
    target = db.Column(db.String(180))
    details = db.Column(db.String(500))


class BackupRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    filename = db.Column(db.String(255), nullable=False, unique=True)
    path = db.Column(db.String(500), nullable=False)
    backup_type = db.Column(db.String(30), default="manual", nullable=False, index=True)
    status = db.Column(db.String(30), default="RUNNING", nullable=False, index=True)
    created_by = db.Column(db.String(120))
    auth_method = db.Column(db.String(30))
    sso_groups = db.Column(db.Text)
    file_count = db.Column(db.Integer, default=0, nullable=False)
    size_bytes = db.Column(db.Integer, default=0, nullable=False)
    sha256 = db.Column(db.String(64))
    manifest_json = db.Column(db.Text)
    verified_at = db.Column(db.DateTime)
    verify_message = db.Column(db.String(255))


def station_label(code, short=False):
    station = WORKSTATION_BY_CODE.get(code)
    if not station:
        return code
    return station["short"] if short else station["name"]


def station_group(code):
    return f"{SSO_STATION_GROUP_PREFIX}/{code}"


def actor_username():
    try:
        if current_user.is_authenticated:
            return current_user.username
    except RuntimeError:
        pass
    return "System"


def log_action(action, target, details=""):
    try:
        db.session.add(
            AuditLog(
                username=actor_username(),
                action=action,
                target=target,
                details=(details or "")[:500],
            )
        )
        db.session.commit()
    except Exception:
        db.session.rollback()


def parse_sso_groups(value):
    groups = []
    for raw in (value or "").replace(";", ",").split(","):
        group = raw.strip()
        if group:
            groups.append(group)
    return groups


def normalize_remote_address(value):
    address = (value or "").strip()
    if address.startswith("::ffff:"):
        return address[7:]
    return address


def request_peer_address():
    proxy_fix_original = request.environ.get("werkzeug.proxy_fix.orig") or {}
    return normalize_remote_address(proxy_fix_original.get("REMOTE_ADDR") or request.remote_addr or "")


def is_trusted_sso_proxy():
    remote_address = request_peer_address()
    if not remote_address:
        return False
    if remote_address == "::1":
        return True
    try:
        peer = ipaddress.ip_address(remote_address)
    except ValueError:
        return False
    for cidr in SSO_TRUSTED_PROXY_CIDRS:
        try:
            if peer in ipaddress.ip_network(cidr, strict=False):
                return True
        except ValueError:
            continue
    return False


def current_sso_user():
    if not TRUST_SSO_HEADERS or not is_trusted_sso_proxy():
        return None
    groups = parse_sso_groups(request.headers.get("X-SSO-Groups", ""))
    sso_admin = SSO_APP_ADMIN_GROUP in groups or SSO_GLOBAL_ADMIN_GROUP in groups
    sso_user = SSO_APP_USER_GROUP in groups
    has_station_group = any(station_group(code) in groups for code in ALL_STATION_CODES)
    if not (sso_user or sso_admin or has_station_group):
        return None
    username = (
        request.headers.get("X-SSO-Preferred-Username")
        or request.headers.get("X-SSO-User")
        or request.headers.get("X-SSO-Email")
        or ""
    ).strip()
    if not username:
        return None
    return SSOUser(
        username,
        sso_admin=sso_admin,
        groups=groups,
        email=(request.headers.get("X-SSO-Email") or "").strip() or None,
    )


def sync_sso_session():
    sso_user = current_sso_user()
    if not sso_user:
        if (
            TRUST_SSO_HEADERS
            and current_user.is_authenticated
            and getattr(current_user, "auth_method", "local") == "sso"
        ):
            session.pop("sso_user", None)
            logout_user()
        return None

    existing = session.get("sso_user") or {}
    if (
        current_user.is_authenticated
        and getattr(current_user, "auth_method", "local") == "sso"
        and existing.get("id") == sso_user.id
        and existing.get("sso_admin") == sso_user.sso_admin
        and existing.get("sso_groups") == sso_user.sso_groups
    ):
        return current_user

    session["sso_user"] = {
        "id": sso_user.id,
        "username": sso_user.username,
        "sso_admin": sso_user.sso_admin,
        "sso_groups": sso_user.sso_groups,
        "email": sso_user.email,
    }
    login_user(sso_user)
    log_action("LOGIN", "SSO", f"SSO login: {sso_user.username}")
    return sso_user


@login_manager.user_loader
def load_user(user_id):
    if str(user_id).startswith("sso:"):
        sso_session = session.get("sso_user") or {}
        if sso_session.get("id") != user_id:
            return None
        return SSOUser(
            sso_session.get("username", ""),
            sso_admin=sso_session.get("sso_admin", False),
            groups=sso_session.get("sso_groups", []),
            email=sso_session.get("email"),
        )
    try:
        return User.query.get(int(user_id))
    except (TypeError, ValueError):
        return None


@app.before_request
def manage_session():
    if request.endpoint == "static":
        return None
    session.permanent = True
    sync_sso_session()
    if "last_active" in session:
        try:
            last_active = datetime.fromisoformat(session["last_active"])
            if datetime.utcnow() - last_active > app.config["PERMANENT_SESSION_LIFETIME"]:
                session.clear()
                flash("Η συνεδρία έληξε λόγω αδράνειας.", "warning")
                return redirect(url_for("login"))
        except ValueError:
            session.pop("last_active", None)
    session["last_active"] = datetime.utcnow().isoformat()
    return None


@app.after_request
def add_security_headers(response):
    if "text/html" in response.headers.get("Content-Type", ""):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "SAMEORIGIN"
    return response


@app.errorhandler(CSRFError)
def handle_csrf_error(_error):
    flash("Η φόρμα έληξε. Δοκιμάστε ξανά.", "warning")
    return redirect(request.referrer or url_for("index"))


def is_admin_user():
    return current_user.is_authenticated and (
        bool(getattr(current_user, "sso_admin", False))
        or bool(getattr(current_user, "is_admin", False))
        or current_user.username == os.getenv("ADMIN_USERNAME", "admin")
    )


def local_user_admin_allowed():
    if not is_admin_user():
        return False
    if getattr(current_user, "auth_method", "local") == "sso":
        return ALLOW_LOCAL_USER_ADMIN_FROM_SSO
    return True


def allowed_station_codes():
    if not current_user.is_authenticated:
        return set()
    if is_admin_user():
        return set(ALL_STATION_CODES)
    if getattr(current_user, "auth_method", "local") == "sso":
        groups = set(getattr(current_user, "sso_groups", []) or [])
        station_codes = {code for code in ALL_STATION_CODES if station_group(code) in groups}
        if ALLOW_SSO_USER_GROUP_ALL_STATIONS and SSO_APP_USER_GROUP in groups:
            station_codes.update(ALL_STATION_CODES)
        return station_codes
    return {
        permission.station_code
        for permission in LocalUserStationPermission.query.filter_by(user_id=current_user.id).all()
    }


def can_complete_step(step):
    if not current_user.is_authenticated or step.case.status == "cancelled":
        return False
    if is_admin_user():
        return True
    if step.station_code not in allowed_station_codes():
        return False
    active_step = current_step_for_case(step.case)
    return active_step is not None and active_step.id == step.id


def can_update_step(step):
    if not current_user.is_authenticated or step.case.status == "cancelled":
        return False
    if is_admin_user():
        return True
    return step.station_code in allowed_station_codes()


def pending_action_summary(limit=5):
    summary = {"count": 0, "cases": [], "station_codes": []}
    if not current_user.is_authenticated:
        return summary
    station_codes = sorted(allowed_station_codes())
    if not station_codes:
        return summary
    summary["station_codes"] = station_codes
    open_cases = SurgeryCase.query.filter_by(status="open").order_by(SurgeryCase.updated_at.desc()).all()
    for case in open_cases:
        active_step = current_step_for_case(case)
        if active_step and active_step.station_code in station_codes:
            summary["count"] += 1
            if len(summary["cases"]) < limit:
                summary["cases"].append(case)
    return summary


@app.context_processor
def inject_template_globals():
    return {
        "is_admin_user": is_admin_user,
        "allowed_station_codes": allowed_station_codes,
        "can_complete_step": can_complete_step,
        "can_update_step": can_update_step,
        "station_label": station_label,
        "station_group": station_group,
        "workstations": WORKSTATIONS,
        "case_status_labels": CASE_STATUS_LABELS,
        "step_status_labels": STEP_STATUS_LABELS,
        "delay_warning_days": DELAY_WARNING_DAYS,
        "current_step_for_case": current_step_for_case,
        "current_step_age_days": current_step_age_days,
        "progress_for_case": progress_for_case,
        "is_case_delayed": is_case_delayed,
        "pending_action_summary": pending_action_summary,
    }


@app.template_filter("datetime_el")
def datetime_el(value):
    if not value:
        return "-"
    return value.strftime("%d/%m/%Y %H:%M")


@app.template_filter("date_el")
def date_el(value):
    if not value:
        return "-"
    return value.strftime("%d/%m/%Y")


@app.template_filter("filesize")
def filesize_filter(value):
    try:
        size = float(value or 0)
    except (TypeError, ValueError):
        return "0 B"
    for unit in ["B", "KB", "MB", "GB", "TB"]:
        if size < 1024 or unit == "TB":
            return f"{int(size)} {unit}" if unit == "B" else f"{size:.1f} {unit}"
        size /= 1024
    return "0 B"


def create_all_with_retry(retries=30, delay_seconds=1):
    for attempt in range(retries):
        try:
            db.create_all()
            return
        except Exception:
            db.session.rollback()
            if attempt == retries - 1:
                raise
            time.sleep(delay_seconds)


def seed_admin_user():
    admin_username = os.getenv("ADMIN_USERNAME", "admin")
    admin_password = os.getenv("ADMIN_PASSWORD", "admin12345")
    admin = User.query.filter_by(username=admin_username).first()
    if admin:
        if not admin.is_admin:
            admin.is_admin = True
            db.session.commit()
        return
    db.session.add(
        User(
            username=admin_username,
            password_hash=generate_password_hash(admin_password),
            is_admin=True,
        )
    )
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()


with app.app_context():
    create_all_with_retry()
    seed_admin_user()


def generate_reference_code(case_id):
    return f"MMFI-{case_id:06d}"


def current_step_for_case(case):
    for step in sorted(case.steps, key=lambda item: item.step_order):
        if step.status != "completed":
            return step
    return None


def progress_for_case(case):
    total = len(case.steps)
    if total == 0:
        return 0
    done = len([step for step in case.steps if step.status == "completed"])
    return round(done * 100 / total)


def current_step_age_days(case):
    step = current_step_for_case(case)
    if not step:
        return 0
    anchor = None
    previous = [item for item in case.steps if item.step_order < step.step_order and item.completed_at]
    if previous:
        anchor = max(item.completed_at for item in previous)
    else:
        anchor = datetime.combine(case.surgery_date, datetime.min.time())
    return max((datetime.utcnow() - anchor).days, 0)


def is_case_delayed(case):
    return case.status == "open" and current_step_age_days(case) >= DELAY_WARNING_DAYS


def create_workflow_steps(case):
    for item in WORKFLOW_DEFINITION:
        station = WORKSTATION_BY_CODE[item["station_code"]]
        db.session.add(
            WorkflowStep(
                case=case,
                step_order=item["order"],
                station_code=item["station_code"],
                station_name=station["name"],
                action_title=item["title"],
                action_description=item["description"],
                status="pending",
            )
        )


def update_case_rollup(case):
    if case.status == "cancelled":
        return
    if case.steps and all(step.status == "completed" for step in case.steps):
        case.status = "completed"
        completed_dates = [step.completed_at for step in case.steps if step.completed_at]
        case.completed_at = max(completed_dates) if completed_dates else datetime.utcnow()
    else:
        case.status = "open"
        case.completed_at = None
    case.updated_at = datetime.utcnow()


def parse_date(value, field_name="ημερομηνία"):
    if not value:
        raise ValueError(f"Συμπληρώστε {field_name}.")
    return datetime.strptime(value, "%Y-%m-%d").date()


def filtered_cases_query():
    query = SurgeryCase.query
    status = request.args.get("status", "").strip()
    station = request.args.get("station", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()
    q = request.args.get("q", "").strip()

    if status:
        query = query.filter(SurgeryCase.status == status)
    if date_from:
        query = query.filter(SurgeryCase.surgery_date >= parse_date(date_from, "από"))
    if date_to:
        query = query.filter(SurgeryCase.surgery_date <= parse_date(date_to, "έως"))
    if q:
        like = f"%{q}%"
        query = query.filter(
            db.or_(
                SurgeryCase.reference_code.ilike(like),
                SurgeryCase.surgery_description.ilike(like),
                SurgeryCase.responsible_doctor.ilike(like),
                SurgeryCase.supplier.ilike(like),
                SurgeryCase.internal_reference.ilike(like),
            )
        )
    if station:
        active_ids = []
        for case in query.order_by(SurgeryCase.surgery_date.desc()).all():
            active = current_step_for_case(case)
            if active and active.station_code == station:
                active_ids.append(case.id)
        query = SurgeryCase.query.filter(SurgeryCase.id.in_(active_ids or [-1]))
    return query


def cases_for_current_filters():
    try:
        return filtered_cases_query().order_by(SurgeryCase.surgery_date.desc(), SurgeryCase.id.desc()).all()
    except ValueError as exc:
        flash(str(exc), "warning")
        return SurgeryCase.query.order_by(SurgeryCase.surgery_date.desc(), SurgeryCase.id.desc()).all()


def dashboard_context():
    open_cases = SurgeryCase.query.filter_by(status="open").order_by(SurgeryCase.updated_at.desc()).all()
    completed_count = SurgeryCase.query.filter_by(status="completed").count()
    all_count = SurgeryCase.query.count()
    delayed_cases = [case for case in open_cases if is_case_delayed(case)]
    my_station_codes = allowed_station_codes()
    my_tasks = [
        case
        for case in open_cases
        if (current_step_for_case(case) and current_step_for_case(case).station_code in my_station_codes)
    ]
    return {
        "open_cases": open_cases,
        "completed_count": completed_count,
        "all_count": all_count,
        "delayed_cases": delayed_cases,
        "my_tasks": my_tasks[:12],
        "recent_cases": SurgeryCase.query.order_by(SurgeryCase.updated_at.desc()).limit(12).all(),
    }


@app.route("/health")
def health():
    return {"status": "ok", "app": "mmfi", "time": datetime.utcnow().isoformat(timespec="seconds") + "Z"}


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = User.query.filter_by(username=username, is_active_local=True).first()
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            log_action("LOGIN", "Local", f"Local login: {username}")
            return redirect(url_for("index"))
        log_action("FAILED_LOGIN", "Local", username)
        flash("Λάθος όνομα χρήστη ή κωδικός.", "danger")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    log_action("LOGOUT", "Session", "User logout")
    session.pop("sso_user", None)
    logout_user()
    return redirect(url_for("login"))


@app.route("/change_password", methods=["GET", "POST"])
@login_required
def change_password():
    if getattr(current_user, "auth_method", "local") == "sso":
        flash("Η αλλαγή κωδικού γίνεται από το κεντρικό Auth.", "warning")
        return redirect(url_for("index"))
    if request.method == "POST":
        old_password = request.form.get("old_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")
        user = User.query.get_or_404(current_user.id)
        if not check_password_hash(user.password_hash, old_password):
            flash("Ο παλιός κωδικός δεν είναι σωστός.", "danger")
            return redirect(url_for("change_password"))
        if len(new_password) < 8:
            flash("Ο νέος κωδικός πρέπει να έχει τουλάχιστον 8 χαρακτήρες.", "warning")
            return redirect(url_for("change_password"))
        if new_password != confirm_password:
            flash("Οι νέοι κωδικοί δεν ταιριάζουν.", "warning")
            return redirect(url_for("change_password"))
        user.password_hash = generate_password_hash(new_password)
        db.session.commit()
        log_action("PASSWORD_CHANGE", f"User:{user.username}", "Password changed")
        flash("Ο κωδικός άλλαξε.", "success")
        return redirect(url_for("index"))
    return render_template("change_password.html")


@app.route("/")
@login_required
def index():
    return render_template("index.html", **dashboard_context())


@app.route("/cases")
@login_required
def cases():
    selected_cases = cases_for_current_filters()
    return render_template("cases.html", cases=selected_cases)


@app.route("/cases/new", methods=["GET", "POST"])
@login_required
def create_case():
    if request.method == "POST":
        try:
            surgery_date = parse_date(request.form.get("surgery_date"), "ημερομηνία χειρουργείου")
            description = request.form.get("surgery_description", "").strip()
            responsible_doctor = request.form.get("responsible_doctor", "").strip()
            if not description or not responsible_doctor:
                raise ValueError("Συμπληρώστε περιγραφή χειρουργείου και υπεύθυνο ιατρό.")
            case = SurgeryCase(
                reference_code=f"PENDING-{uuid.uuid4().hex[:10]}",
                surgery_date=surgery_date,
                surgery_description=description,
                responsible_doctor=responsible_doctor,
                supplier=request.form.get("supplier", "").strip() or None,
                internal_reference=request.form.get("internal_reference", "").strip() or None,
                notes=request.form.get("notes", "").strip() or None,
                priority=request.form.get("priority", "normal"),
                created_by=actor_username(),
            )
            db.session.add(case)
            db.session.flush()
            case.reference_code = generate_reference_code(case.id)
            create_workflow_steps(case)
            db.session.commit()
            log_action("CASE_CREATE", case.reference_code, "New surgery financial tracking case")
            flash(f"Δημιουργήθηκε η υπόθεση {case.reference_code}.", "success")
            return redirect(url_for("case_detail", case_id=case.id))
        except ValueError as exc:
            flash(str(exc), "warning")
        except Exception as exc:
            db.session.rollback()
            flash(f"Δεν ήταν δυνατή η δημιουργία: {exc}", "danger")
    return render_template("case_form.html", case=None)


@app.route("/cases/<int:case_id>")
@login_required
def case_detail(case_id):
    case = SurgeryCase.query.get_or_404(case_id)
    return render_template(
        "case_detail.html",
        case=case,
        current_step=current_step_for_case(case),
        progress=progress_for_case(case),
        is_delayed=is_case_delayed(case),
        step_age_days=current_step_age_days(case),
    )


@app.route("/cases/<int:case_id>/edit", methods=["GET", "POST"])
@login_required
def edit_case(case_id):
    case = SurgeryCase.query.get_or_404(case_id)
    if not is_admin_user() and actor_username() != case.created_by:
        flash("Δεν έχετε δικαίωμα επεξεργασίας αυτής της υπόθεσης.", "danger")
        return redirect(url_for("case_detail", case_id=case.id))
    if request.method == "POST":
        try:
            case.surgery_date = parse_date(request.form.get("surgery_date"), "ημερομηνία χειρουργείου")
            case.surgery_description = request.form.get("surgery_description", "").strip()
            case.responsible_doctor = request.form.get("responsible_doctor", "").strip()
            if not case.surgery_description or not case.responsible_doctor:
                raise ValueError("Συμπληρώστε περιγραφή χειρουργείου και υπεύθυνο ιατρό.")
            case.supplier = request.form.get("supplier", "").strip() or None
            case.internal_reference = request.form.get("internal_reference", "").strip() or None
            case.notes = request.form.get("notes", "").strip() or None
            case.priority = request.form.get("priority", "normal")
            case.updated_at = datetime.utcnow()
            db.session.commit()
            log_action("CASE_UPDATE", case.reference_code, "Case edited")
            flash("Οι αλλαγές αποθηκεύτηκαν.", "success")
            return redirect(url_for("case_detail", case_id=case.id))
        except ValueError as exc:
            flash(str(exc), "warning")
        except Exception as exc:
            db.session.rollback()
            flash(f"Δεν ήταν δυνατή η αποθήκευση: {exc}", "danger")
    return render_template("case_form.html", case=case)


@app.route("/cases/<int:case_id>/cancel", methods=["POST"])
@login_required
def cancel_case(case_id):
    if not is_admin_user():
        flash("Μόνο διαχειριστής μπορεί να ακυρώσει υπόθεση.", "danger")
        return redirect(url_for("case_detail", case_id=case_id))
    case = SurgeryCase.query.get_or_404(case_id)
    case.status = "cancelled"
    case.cancelled_at = datetime.utcnow()
    case.cancelled_by = actor_username()
    case.cancellation_reason = request.form.get("cancellation_reason", "").strip() or None
    db.session.commit()
    log_action("CASE_CANCEL", case.reference_code, case.cancellation_reason or "")
    flash("Η υπόθεση ακυρώθηκε.", "warning")
    return redirect(url_for("case_detail", case_id=case.id))


@app.route("/cases/<int:case_id>/reopen", methods=["POST"])
@login_required
def reopen_case(case_id):
    if not is_admin_user():
        flash("Μόνο διαχειριστής μπορεί να επαναφέρει υπόθεση.", "danger")
        return redirect(url_for("case_detail", case_id=case_id))
    case = SurgeryCase.query.get_or_404(case_id)
    case.status = "open"
    case.cancelled_at = None
    case.cancelled_by = None
    case.cancellation_reason = None
    update_case_rollup(case)
    db.session.commit()
    log_action("CASE_REOPEN", case.reference_code, "Case reopened")
    flash("Η υπόθεση επανήλθε σε εξέλιξη.", "success")
    return redirect(url_for("case_detail", case_id=case.id))


@app.route("/cases/<int:case_id>/steps/<int:step_id>/complete", methods=["POST"])
@login_required
def complete_step(case_id, step_id):
    case = SurgeryCase.query.get_or_404(case_id)
    step = WorkflowStep.query.filter_by(id=step_id, case_id=case.id).first_or_404()
    if step.status == "completed":
        flash("Το βήμα έχει ήδη ολοκληρωθεί.", "info")
        return redirect(url_for("case_detail", case_id=case.id))
    if not can_complete_step(step):
        flash("Δεν έχετε δικαίωμα ολοκλήρωσης για αυτό το βήμα.", "danger")
        return redirect(url_for("case_detail", case_id=case.id))
    if not is_admin_user():
        active_step = current_step_for_case(case)
        if not active_step or active_step.id != step.id:
            flash("Μπορεί να ολοκληρωθεί μόνο το τρέχον ενεργό βήμα.", "warning")
            return redirect(url_for("case_detail", case_id=case.id))
    try:
        comment = request.form.get("comment", "").strip() or request.form.get("notes", "").strip()
        create_step_history_entry(
            case,
            step,
            event_type="completion",
            comment=comment,
            files=request.files.getlist("history_files"),
        )
        step.status = "completed"
        step.completed_at = datetime.utcnow()
        step.completed_by = actor_username()
        update_case_rollup(case)
        db.session.commit()
        log_action("STEP_COMPLETE", case.reference_code, f"Step {step.step_order}: {step.action_title}")
        flash("Η απάντηση κλειδώθηκε στο ιστορικό και η ενέργεια ολοκληρώθηκε.", "success")
    except ValueError as exc:
        db.session.rollback()
        flash(str(exc), "danger")
    return redirect(url_for("case_detail", case_id=case.id))


@app.route("/cases/<int:case_id>/steps/<int:step_id>/updates", methods=["POST"])
@login_required
def add_step_update(case_id, step_id):
    case = SurgeryCase.query.get_or_404(case_id)
    step = WorkflowStep.query.filter_by(id=step_id, case_id=case.id).first_or_404()
    if step.status != "completed":
        flash("Update γίνεται μόνο σε ήδη ολοκληρωμένο βήμα.", "warning")
        return redirect(url_for("case_detail", case_id=case.id))
    if not can_update_step(step):
        flash("Δεν έχετε δικαίωμα update για αυτό το βήμα.", "danger")
        return redirect(url_for("case_detail", case_id=case.id))
    comment = request.form.get("comment", "").strip()
    files = request.files.getlist("history_files")
    if not comment and not non_empty_uploads(files):
        flash("Συμπληρώστε σχόλιο ή ανεβάστε αρχείο για να δημιουργηθεί update.", "warning")
        return redirect(url_for("case_detail", case_id=case.id))
    try:
        entry = create_step_history_entry(
            case,
            step,
            event_type="update",
            comment=comment,
            files=files,
        )
        db.session.commit()
        log_action("STEP_UPDATE", case.reference_code, f"Step {step.step_order}, version {entry.version_number}")
        flash("Το update προστέθηκε στο ιστορικό χωρίς να χαθεί η προηγούμενη απάντηση.", "success")
    except ValueError as exc:
        db.session.rollback()
        flash(str(exc), "danger")
    return redirect(url_for("case_detail", case_id=case.id))


@app.route("/cases/<int:case_id>/steps/<int:step_id>/reopen", methods=["POST"])
@login_required
def reopen_step(case_id, step_id):
    if not is_admin_user():
        flash("Μόνο διαχειριστής μπορεί να ανοίξει ξανά βήμα.", "danger")
        return redirect(url_for("case_detail", case_id=case_id))
    case = SurgeryCase.query.get_or_404(case_id)
    step = WorkflowStep.query.filter_by(id=step_id, case_id=case.id).first_or_404()
    step.status = "pending"
    step.completed_at = None
    step.completed_by = None
    step.notes = request.form.get("notes", "").strip() or step.notes
    for later_step in case.steps:
        if later_step.step_order > step.step_order:
            later_step.status = "pending"
            later_step.completed_at = None
            later_step.completed_by = None
    update_case_rollup(case)
    db.session.commit()
    log_action("STEP_REOPEN", case.reference_code, f"Step {step.step_order}")
    flash("Το βήμα άνοιξε ξανά και τα επόμενα βήματα επανήλθαν σε εκκρεμότητα.", "warning")
    return redirect(url_for("case_detail", case_id=case.id))


def allowed_upload(filename):
    allowed = {"pdf", "png", "jpg", "jpeg", "doc", "docx", "xls", "xlsx", "txt"}
    return "." in filename and filename.rsplit(".", 1)[1].lower() in allowed


def step_history_next_version(step):
    current_max = (
        db.session.query(db.func.max(StepHistoryEntry.version_number))
        .filter_by(step_id=step.id)
        .scalar()
    )
    return (current_max or 0) + 1


def non_empty_uploads(files):
    return [file for file in files if file and file.filename]


def create_step_history_entry(case, step, *, event_type, comment, files=None):
    upload_files = non_empty_uploads(files or [])
    if upload_files:
        for file in upload_files:
            if not allowed_upload(file.filename):
                raise ValueError(f"Μη αποδεκτός τύπος αρχείου: {file.filename}")

    entry = StepHistoryEntry(
        case=case,
        step=step,
        version_number=step_history_next_version(step),
        event_type=event_type,
        comment=(comment or "").strip() or None,
        created_by=actor_username(),
    )
    db.session.add(entry)
    db.session.flush()

    if entry.comment:
        step.notes = entry.comment

    if upload_files:
        case_folder = os.path.join(app.config["UPLOAD_FOLDER"], case.reference_code, f"step_{step.step_order}")
        os.makedirs(case_folder, exist_ok=True)
        for file in upload_files:
            original = secure_filename(file.filename) or "attachment"
            stored = f"history_{entry.id}_{uuid.uuid4().hex}_{original}"
            destination = os.path.join(case_folder, stored)
            file.save(destination)
            relative_path = os.path.relpath(destination, BASE_DIR).replace("\\", "/")
            db.session.add(
                StepHistoryAttachment(
                    history_entry=entry,
                    case=case,
                    step=step,
                    original_filename=file.filename,
                    stored_filename=stored,
                    relative_path=relative_path,
                    uploaded_by=actor_username(),
                )
            )

    case.updated_at = datetime.utcnow()
    return entry


@app.route("/cases/<int:case_id>/attachments", methods=["POST"])
@login_required
def upload_attachment(case_id):
    case = SurgeryCase.query.get_or_404(case_id)
    file = request.files.get("attachment")
    if not file or not file.filename:
        flash("Επιλέξτε αρχείο.", "warning")
        return redirect(url_for("case_detail", case_id=case.id))
    if not allowed_upload(file.filename):
        flash("Μη αποδεκτός τύπος αρχείου.", "danger")
        return redirect(url_for("case_detail", case_id=case.id))
    original = secure_filename(file.filename) or "attachment"
    stored = f"{case.reference_code}_{uuid.uuid4().hex}_{original}"
    case_folder = os.path.join(app.config["UPLOAD_FOLDER"], case.reference_code)
    os.makedirs(case_folder, exist_ok=True)
    destination = os.path.join(case_folder, stored)
    file.save(destination)
    relative_path = os.path.relpath(destination, BASE_DIR).replace("\\", "/")
    attachment = CaseAttachment(
        case=case,
        step=None,
        original_filename=file.filename,
        stored_filename=stored,
        relative_path=relative_path,
        description=request.form.get("description", "").strip() or None,
        uploaded_by=actor_username(),
    )
    db.session.add(attachment)
    db.session.commit()
    log_action("ATTACHMENT_UPLOAD", case.reference_code, file.filename)
    flash("Το αρχείο ανέβηκε.", "success")
    return redirect(url_for("case_detail", case_id=case.id))


@app.route("/attachments/<int:attachment_id>/download")
@login_required
def download_attachment(attachment_id):
    attachment = CaseAttachment.query.get_or_404(attachment_id)
    path = os.path.join(BASE_DIR, attachment.relative_path)
    if not os.path.exists(path):
        flash("Το αρχείο δεν βρέθηκε στον δίσκο.", "danger")
        return redirect(url_for("case_detail", case_id=attachment.case_id))
    log_action("ATTACHMENT_DOWNLOAD", attachment.case.reference_code, attachment.original_filename)
    return send_file(path, as_attachment=True, download_name=attachment.original_filename)


@app.route("/step-history-files/<int:file_id>/download")
@login_required
def download_step_history_file(file_id):
    history_file = StepHistoryAttachment.query.get_or_404(file_id)
    path = os.path.join(BASE_DIR, history_file.relative_path)
    if not os.path.exists(path):
        flash("Το ιστορικό αρχείο δεν βρέθηκε στον δίσκο.", "danger")
        return redirect(url_for("case_detail", case_id=history_file.case_id))
    log_action(
        "STEP_HISTORY_FILE_DOWNLOAD",
        history_file.case.reference_code,
        history_file.original_filename,
    )
    return send_file(path, as_attachment=True, download_name=history_file.original_filename)


@app.route("/attachments/<int:attachment_id>/delete", methods=["POST"])
@login_required
def delete_attachment(attachment_id):
    if not is_admin_user():
        flash("Μόνο διαχειριστής μπορεί να διαγράψει συνημμένο.", "danger")
        return redirect(request.referrer or url_for("index"))
    attachment = CaseAttachment.query.get_or_404(attachment_id)
    if attachment.step_id:
        flash("Συνημμένα που έχουν συνδεθεί με βήμα δεν διαγράφονται από την εφαρμογή.", "warning")
        return redirect(url_for("case_detail", case_id=attachment.case_id))
    case_id = attachment.case_id
    path = os.path.join(BASE_DIR, attachment.relative_path)
    if os.path.exists(path):
        os.remove(path)
    log_action("ATTACHMENT_DELETE", attachment.case.reference_code, attachment.original_filename)
    db.session.delete(attachment)
    db.session.commit()
    flash("Το συνημμένο διαγράφηκε.", "success")
    return redirect(url_for("case_detail", case_id=case_id))


@app.route("/reports")
@login_required
def reports():
    selected_cases = cases_for_current_filters()
    return render_template("reports.html", cases=selected_cases)


def apply_report_headers(sheet, headers):
    sheet.append(headers)
    fill = PatternFill("solid", fgColor="1F4E79")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


@app.route("/reports/export.xlsx")
@login_required
def export_report_xlsx():
    selected_cases = cases_for_current_filters()
    wb = Workbook()
    ws = wb.active
    ws.title = "Υποθέσεις"
    apply_report_headers(
        ws,
        [
            "Κωδικός",
            "Ημ/νία χειρουργείου",
            "Περιγραφή",
            "Υπεύθυνος ιατρός",
            "Προμηθευτής",
            "Κατάσταση",
            "Τρέχον βήμα",
            "Τρέχων σταθμός",
            "Ημέρες στο βήμα",
            "Πρόοδος",
            "Δημιουργήθηκε",
            "Δημιουργός",
        ],
    )
    for case in selected_cases:
        active = current_step_for_case(case)
        ws.append(
            [
                case.reference_code,
                case.surgery_date.strftime("%d/%m/%Y"),
                case.surgery_description,
                case.responsible_doctor,
                case.supplier or "",
                CASE_STATUS_LABELS.get(case.status, case.status),
                active.action_title if active else "",
                active.station_name if active else "",
                current_step_age_days(case) if active else "",
                f"{progress_for_case(case)}%",
                case.created_at.strftime("%d/%m/%Y %H:%M"),
                case.created_by or "",
            ]
        )

    steps_ws = wb.create_sheet("Βήματα")
    apply_report_headers(
        steps_ws,
        [
            "Κωδικός υπόθεσης",
            "Α/Α",
            "Σταθμός",
            "Ενέργεια",
            "Κατάσταση",
            "Ημ/νία ολοκλήρωσης",
            "Χρήστης",
            "Σχόλια",
        ],
    )
    for case in selected_cases:
        for step in case.steps:
            steps_ws.append(
                [
                    case.reference_code,
                    step.step_order,
                    step.station_name,
                    step.action_title,
                    STEP_STATUS_LABELS.get(step.status, step.status),
                    step.completed_at.strftime("%d/%m/%Y %H:%M") if step.completed_at else "",
                    step.completed_by or "",
                    step.notes or "",
                ]
            )

    history_ws = wb.create_sheet("Ιστορικό")
    apply_report_headers(
        history_ws,
        [
            "Κωδικός υπόθεσης",
            "Βήμα",
            "Έκδοση",
            "Τύπος",
            "Χρήστης",
            "Ημερομηνία",
            "Σχόλιο",
            "Αρχεία",
        ],
    )
    for case in selected_cases:
        for step in case.steps:
            for entry in sorted(step.history_entries, key=lambda item: item.version_number):
                history_ws.append(
                    [
                        case.reference_code,
                        step.step_order,
                        entry.version_number,
                        entry.event_type,
                        entry.created_by or "",
                        entry.created_at.strftime("%d/%m/%Y %H:%M") if entry.created_at else "",
                        entry.comment or "",
                        ", ".join(file.original_filename for file in entry.files),
                    ]
                )

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 55)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    wb.save(tmp_path)
    log_action("REPORT_EXPORT", "XLSX", f"{len(selected_cases)} cases")
    return send_file(
        tmp_path,
        as_attachment=True,
        download_name=f"mmfi_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
    )


@app.route("/audit")
@login_required
def audit():
    if not is_admin_user():
        flash("Δεν έχετε δικαίωμα πρόσβασης.", "danger")
        return redirect(url_for("index"))
    logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).limit(300).all()
    return render_template("audit.html", logs=logs)


def central_auth_context():
    groups = [
        {
            "name": SSO_APP_USER_GROUP,
            "role": "Πρόσβαση εφαρμογής",
            "description": "Είσοδος στην εφαρμογή MMFI.",
        },
        {
            "name": SSO_APP_ADMIN_GROUP,
            "role": "Διαχείριση εφαρμογής",
            "description": "Πλήρης διαχείριση υποθέσεων, χρηστών, backup και audit.",
        },
        {
            "name": SSO_GLOBAL_ADMIN_GROUP,
            "role": "Global admin",
            "description": "Κοινός διαχειριστής όλων των intranet εφαρμογών.",
        },
    ]
    for station in WORKSTATIONS:
        groups.append(
            {
                "name": station_group(station["code"]),
                "role": station["name"],
                "description": "Δικαίωμα ολοκλήρωσης ενεργειών του συγκεκριμένου σταθμού.",
            }
        )
    return {
        "realm": CENTRAL_AUTH_REALM,
        "admin_url": CENTRAL_AUTH_ADMIN_URL,
        "users_url": CENTRAL_AUTH_USERS_URL,
        "groups_url": CENTRAL_AUTH_GROUPS_URL,
        "groups": groups,
        "current_groups": getattr(current_user, "sso_groups", []),
        "auth_method": getattr(current_user, "auth_method", "local"),
    }


@app.route("/manage_users")
@login_required
def manage_users():
    if not is_admin_user():
        flash("Δεν έχετε δικαίωμα πρόσβασης.", "danger")
        return redirect(url_for("index"))
    users = User.query.order_by(User.username.asc()).all()
    permissions = {}
    for permission in LocalUserStationPermission.query.all():
        permissions.setdefault(permission.user_id, set()).add(permission.station_code)
    return render_template(
        "manage_users.html",
        users=users,
        permissions=permissions,
        central_auth=central_auth_context(),
        local_user_admin_allowed=local_user_admin_allowed(),
        admin_username=os.getenv("ADMIN_USERNAME", "admin"),
    )


@app.route("/users/create", methods=["POST"])
@login_required
def create_user():
    if not local_user_admin_allowed():
        flash("Η τοπική διαχείριση χρηστών δεν είναι διαθέσιμη από SSO.", "danger")
        return redirect(url_for("manage_users"))
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")
    if not username or len(password) < 8:
        flash("Συμπληρώστε username και κωδικό τουλάχιστον 8 χαρακτήρων.", "warning")
        return redirect(url_for("manage_users"))
    user = User(
        username=username,
        password_hash=generate_password_hash(password),
        is_admin=bool(request.form.get("is_admin")),
    )
    db.session.add(user)
    try:
        db.session.flush()
        for code in request.form.getlist("station_codes"):
            if code in ALL_STATION_CODES:
                db.session.add(LocalUserStationPermission(user=user, station_code=code))
        db.session.commit()
        log_action("USER_CREATE", username, "Local fallback user created")
        flash("Ο τοπικός χρήστης δημιουργήθηκε.", "success")
    except IntegrityError:
        db.session.rollback()
        flash("Υπάρχει ήδη χρήστης με αυτό το username.", "danger")
    return redirect(url_for("manage_users"))


@app.route("/users/<int:user_id>/permissions", methods=["POST"])
@login_required
def update_user_permissions(user_id):
    if not local_user_admin_allowed():
        flash("Η τοπική διαχείριση χρηστών δεν είναι διαθέσιμη από SSO.", "danger")
        return redirect(url_for("manage_users"))
    user = User.query.get_or_404(user_id)
    user.is_admin = bool(request.form.get("is_admin"))
    user.is_active_local = bool(request.form.get("is_active_local"))
    LocalUserStationPermission.query.filter_by(user_id=user.id).delete()
    for code in request.form.getlist("station_codes"):
        if code in ALL_STATION_CODES:
            db.session.add(LocalUserStationPermission(user=user, station_code=code))
    new_password = request.form.get("new_password", "")
    if new_password:
        if len(new_password) < 8:
            flash("Ο νέος κωδικός πρέπει να έχει τουλάχιστον 8 χαρακτήρες.", "warning")
            return redirect(url_for("manage_users"))
        user.password_hash = generate_password_hash(new_password)
    db.session.commit()
    log_action("USER_UPDATE", user.username, "Local fallback user updated")
    flash("Τα δικαιώματα ενημερώθηκαν.", "success")
    return redirect(url_for("manage_users"))


@app.route("/users/<int:user_id>/delete", methods=["POST"])
@login_required
def delete_user(user_id):
    if not local_user_admin_allowed():
        flash("Η τοπική διαχείριση χρηστών δεν είναι διαθέσιμη από SSO.", "danger")
        return redirect(url_for("manage_users"))
    user = User.query.get_or_404(user_id)
    admin_username = os.getenv("ADMIN_USERNAME", "admin")
    if user.username == admin_username:
        flash("Ο βασικός admin fallback δεν διαγράφεται.", "warning")
        return redirect(url_for("manage_users"))
    log_action("USER_DELETE", user.username, "Local fallback user deleted")
    db.session.delete(user)
    db.session.commit()
    flash("Ο χρήστης διαγράφηκε.", "success")
    return redirect(url_for("manage_users"))


def current_actor_context():
    try:
        if current_user.is_authenticated:
            return {
                "username": current_user.username,
                "auth_method": getattr(current_user, "auth_method", "local"),
                "sso_groups": list(getattr(current_user, "sso_groups", []) or []),
            }
    except RuntimeError:
        pass
    return {"username": "System", "auth_method": "system", "sso_groups": []}


def backup_password():
    return required_env("BACKUP_PASSWORD", "change-me-local-backup-password").encode("utf-8")


def safe_backup_path(filename):
    backup_root = os.path.abspath(BACKUP_FOLDER)
    path = os.path.abspath(os.path.join(backup_root, os.path.basename(filename)))
    if os.path.commonpath([backup_root, path]) != backup_root:
        raise ValueError("Invalid backup path")
    return path


def sha256_file(path):
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def git_revision():
    version = os.getenv("APP_VERSION")
    if version:
        return version
    try:
        result = subprocess.run(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=BASE_DIR,
            capture_output=True,
            text=True,
            timeout=5,
            check=False,
        )
        if result.returncode == 0:
            return result.stdout.strip()
    except Exception:
        pass
    return "unknown"


def iter_directory_files(base_path, archive_prefix, *, exclude_dirs=None):
    exclude_dirs = set(exclude_dirs or [])
    root_path = os.path.abspath(base_path)
    if not os.path.isdir(root_path):
        return
    for root, dirs, files in os.walk(root_path):
        dirs[:] = [item for item in dirs if item not in exclude_dirs and item != "__pycache__"]
        for filename in files:
            file_path = os.path.join(root, filename)
            if not os.path.isfile(file_path):
                continue
            rel = os.path.relpath(file_path, root_path).replace("\\", "/")
            yield file_path, f"{archive_prefix}/{rel}"


def sqlite_database_path():
    uri = app.config["SQLALCHEMY_DATABASE_URI"]
    if not uri.startswith("sqlite:///"):
        return None
    path = uri.replace("sqlite:///", "", 1)
    if not os.path.isabs(path):
        path = os.path.join(BASE_DIR, path)
    return path


def write_database_dump(temp_dir):
    if not BACKUP_INCLUDE_DATABASE:
        note_path = os.path.join(temp_dir, "database_backup_disabled.txt")
        with open(note_path, "w", encoding="utf-8") as handle:
            handle.write("Database backup disabled by BACKUP_INCLUDE_DATABASE=0.\n")
        return note_path, "database/README.txt", "disabled"

    database_url = app.config["SQLALCHEMY_DATABASE_URI"]
    if database_url.startswith("postgresql"):
        pg_dump = shutil.which("pg_dump")
        if not pg_dump:
            raise RuntimeError("pg_dump is not installed in the app container")
        dump_path = os.path.join(temp_dir, "mmfi_database.dump")
        result = subprocess.run(
            [pg_dump, "--dbname", database_url, "--format", "custom", "--file", dump_path],
            capture_output=True,
            text=True,
            timeout=180,
            check=False,
        )
        if result.returncode != 0:
            message = (result.stderr or result.stdout or "pg_dump failed").strip()
            raise RuntimeError(message[:240])
        return dump_path, "database/mmfi_database.dump", "postgresql-custom"

    sqlite_path = sqlite_database_path()
    if sqlite_path and os.path.exists(sqlite_path):
        dump_path = os.path.join(temp_dir, "mmfi.sqlite3")
        shutil.copy2(sqlite_path, dump_path)
        return dump_path, "database/mmfi.sqlite3", "sqlite-copy"

    note_path = os.path.join(temp_dir, "database_not_found.txt")
    with open(note_path, "w", encoding="utf-8") as handle:
        handle.write("No database file or supported DATABASE_URL was found.\n")
    return note_path, "database/README.txt", "missing"


def build_backup_manifest(actor, backup_type, database_mode, entries):
    return {
        "app": "MMFI 251GNA",
        "format_version": 1,
        "created_at_utc": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "created_by": actor["username"],
        "auth_method": actor["auth_method"],
        "sso_groups": actor["sso_groups"],
        "backup_type": backup_type,
        "app_revision": git_revision(),
        "database_mode": database_mode,
        "entries": entries,
    }


def enforce_backup_retention():
    if BACKUP_RETENTION_COUNT <= 0:
        return
    records = (
        BackupRecord.query
        .filter(BackupRecord.status.in_(["CREATED", "VERIFIED"]))
        .order_by(BackupRecord.created_at.desc())
        .all()
    )
    for record in records[BACKUP_RETENTION_COUNT:]:
        try:
            if os.path.exists(record.path):
                os.remove(record.path)
            record.status = "PRUNED"
            record.verify_message = f"Pruned by retention policy: keep last {BACKUP_RETENTION_COUNT}"
        except Exception as exc:
            record.status = "PRUNE_FAILED"
            record.verify_message = str(exc)[:255]
    db.session.commit()


def create_backup_record(backup_type="manual", actor=None):
    actor = actor or current_actor_context()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_filename = f"backup_MMFI_{timestamp}_{backup_type}.zip"
    backup_path = safe_backup_path(backup_filename)
    record = BackupRecord(
        filename=backup_filename,
        path=backup_path,
        backup_type=backup_type,
        status="RUNNING",
        created_by=actor["username"],
        auth_method=actor["auth_method"],
        sso_groups=json.dumps(actor["sso_groups"], ensure_ascii=False),
    )
    db.session.add(record)
    db.session.commit()

    entries = []
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            database_path, database_arcname, database_mode = write_database_dump(temp_dir)
            archive_sources = [(database_path, database_arcname)]
            archive_sources.extend(iter_directory_files(app.config["UPLOAD_FOLDER"], "static/uploads"))
            for _, arcname in archive_sources:
                entries.append(arcname)
            manifest = build_backup_manifest(actor, backup_type, database_mode, entries)
            with pyzipper.AESZipFile(
                backup_path,
                "w",
                compression=pyzipper.ZIP_DEFLATED,
                encryption=pyzipper.WZ_AES,
            ) as archive:
                archive.setpassword(backup_password())
                archive.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))
                for file_path, arcname in archive_sources:
                    archive.write(file_path, arcname)

        record.status = "CREATED"
        record.file_count = len(entries) + 1
        record.size_bytes = os.path.getsize(backup_path)
        record.sha256 = sha256_file(backup_path)
        record.manifest_json = json.dumps(manifest, ensure_ascii=False)
        db.session.commit()
        log_action("BACKUP_CREATE", f"BackupRecord:{record.id}", backup_filename)
        enforce_backup_retention()
        return record
    except Exception as exc:
        db.session.rollback()
        if os.path.exists(backup_path):
            try:
                os.remove(backup_path)
            except OSError:
                pass
        record = BackupRecord.query.get(record.id)
        if record:
            record.status = "FAILED"
            record.verify_message = str(exc)[:255]
            db.session.commit()
        log_action("BACKUP_ERROR", "System", str(exc)[:200])
        raise


def verify_backup_record(record):
    if not record or not os.path.exists(record.path):
        raise FileNotFoundError("Backup archive was not found")
    current_sha256 = sha256_file(record.path)
    if record.sha256 and current_sha256 != record.sha256:
        raise RuntimeError("Backup SHA-256 does not match the stored value")
    with pyzipper.AESZipFile(record.path, "r") as archive:
        archive.setpassword(backup_password())
        names = archive.namelist()
        if "manifest.json" not in names:
            raise RuntimeError("Backup manifest is missing")
        manifest = json.loads(archive.read("manifest.json").decode("utf-8"))
        if manifest.get("app") != "MMFI 251GNA":
            raise RuntimeError("Backup manifest is not for MMFI")
        for name in names:
            archive.read(name)
    record.status = "VERIFIED"
    record.verified_at = datetime.utcnow()
    record.verify_message = f"OK: {len(names)} encrypted entries verified"
    db.session.commit()
    log_action("BACKUP_VERIFY", f"BackupRecord:{record.id}", record.verify_message)
    return True


def backup_system(backup_type="scheduled"):
    with app.app_context():
        try:
            return create_backup_record(backup_type=backup_type, actor=current_actor_context())
        except Exception as exc:
            print(f"[backup] {exc}")
            return None


def start_scheduler_once():
    if os.getenv("DISABLE_SCHEDULER", "0") == "1":
        return None
    scheduler = BackgroundScheduler(daemon=True)
    scheduler.add_job(func=backup_system, trigger="cron", hour=3, minute=0)
    scheduler.start()
    return scheduler


scheduler = start_scheduler_once()


@app.route("/force_backup")
@login_required
def force_backup():
    if not is_admin_user():
        flash("Δεν έχετε δικαίωμα πρόσβασης.", "danger")
        return redirect(url_for("index"))
    record = backup_system(backup_type="manual")
    if record:
        flash(f"Το backup δημιουργήθηκε: {record.filename}", "success")
    else:
        flash("Αποτυχία δημιουργίας backup.", "danger")
    return redirect(url_for("manage_backups"))


@app.route("/manage_backups")
@login_required
def manage_backups():
    if not is_admin_user():
        flash("Δεν έχετε δικαίωμα πρόσβασης.", "danger")
        return redirect(url_for("index"))
    backups = BackupRecord.query.order_by(BackupRecord.created_at.desc()).limit(100).all()
    return render_template(
        "manage_backups.html",
        backups=backups,
        retention_count=BACKUP_RETENTION_COUNT,
        include_database=BACKUP_INCLUDE_DATABASE,
    )


@app.route("/backups/create", methods=["POST"])
@login_required
def create_backup_route():
    if not is_admin_user():
        flash("Δεν έχετε δικαίωμα πρόσβασης.", "danger")
        return redirect(url_for("index"))
    try:
        record = create_backup_record(backup_type="manual")
        flash(f"Το backup δημιουργήθηκε: {record.filename}", "success")
    except Exception as exc:
        flash(f"Αποτυχία δημιουργίας backup: {exc}", "danger")
    return redirect(url_for("manage_backups"))


@app.route("/backups/<int:backup_id>/verify", methods=["POST"])
@login_required
def verify_backup_route(backup_id):
    if not is_admin_user():
        flash("Δεν έχετε δικαίωμα πρόσβασης.", "danger")
        return redirect(url_for("index"))
    record = BackupRecord.query.get_or_404(backup_id)
    try:
        verify_backup_record(record)
        flash("Το backup επαληθεύτηκε.", "success")
    except Exception as exc:
        record.status = "VERIFY_FAILED"
        record.verified_at = datetime.utcnow()
        record.verify_message = str(exc)[:255]
        db.session.commit()
        log_action("BACKUP_VERIFY_ERROR", f"BackupRecord:{record.id}", record.verify_message)
        flash(f"Αποτυχία επαλήθευσης backup: {exc}", "danger")
    return redirect(url_for("manage_backups"))


@app.route("/backups/<int:backup_id>/download")
@login_required
def download_backup_route(backup_id):
    if not is_admin_user():
        flash("Δεν έχετε δικαίωμα πρόσβασης.", "danger")
        return redirect(url_for("index"))
    record = BackupRecord.query.get_or_404(backup_id)
    if not os.path.exists(record.path):
        flash("Το αρχείο backup δεν βρέθηκε.", "danger")
        return redirect(url_for("manage_backups"))
    log_action("BACKUP_DOWNLOAD", f"BackupRecord:{record.id}", record.filename)
    return send_file(record.path, as_attachment=True, download_name=record.filename)


@app.route("/workflow")
@login_required
def workflow():
    return render_template("workflow.html", workflow=WORKFLOW_DEFINITION)


@app.route("/robots.txt")
def robots():
    return Response("User-agent: *\nDisallow: /\n", mimetype="text/plain")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "5050")), debug=bool_env("FLASK_DEBUG", "1"))
